#!/usr/bin/env -S uv run
# /// script
# requires-python = ">=3.11,<3.13"
# dependencies = [
#   "numpy",
#   "openpyxl",
#   "requests",
# ]
# ///
"""Find banter tracks a show is missing and prove where they go by audio.

The original import skipped many taper-labelled "banter" files. For each show
this script:

  1. finds the archive.org item the show was cut from, by matching the item's
     per-file lengths against phish.in's track durations (the catalog's tracks
     are exact butt-cuts of one transfer, so a real match is to the ms);
  2. lists the source files phish.in has no track for, and treats the ones the
     taper labelled banter/talking/etc. as candidates;
  3. downloads just those FLACs (verified against the item's .ffp/.md5), and
     scores the two joints each one would create - tail of the track before it
     spliced to its head, its tail spliced to the head of the track after -
     plus the "direct" joint phish.in currently has between those two tracks.
     If the banter belongs there, both new joints are continuous and the
     direct one is the seam.

Output: <out-dir>/report.json, review.html (listenable joints, approve
checkboxes, approved.json export) and <out-dir>/<date>/*.flac.

Usage (normally via the rake tasks):
  uv run scripts/audio_banter_analysis.py --out-dir data/banter_scan --from-notes
  uv run scripts/audio_banter_analysis.py --out-dir data/banter_scan --dates 1991-05-17,1992-11-20
  uv run scripts/audio_banter_analysis.py --out-dir data/banter_scan --rebuild
"""

import argparse
import base64
import hashlib
import html
import importlib.util
import json
import re
import sys
from concurrent.futures import ThreadPoolExecutor
from dataclasses import asdict, dataclass, field
from pathlib import Path

import numpy as np
import requests

REPO_ROOT = Path(__file__).resolve().parent.parent
JUNCTION_SCRIPT = REPO_ROOT / "scripts" / "audio_junction_analysis.py"
API_BASE = "https://phish.in/api/v2"
SITE_BASE = "https://phish.in"
ARCHIVE = "https://archive.org"
# The Phish Spreadsheet: the fuller source list, for shows archive.org lacks.
SPREADSHEET_ID = "1yAXu83gJBz08cW5OXoqNuN1IbvDXD2vCrDKj4zn1qmU"
SPREADSHEET_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}"
FIRST_YEAR, LAST_YEAR = 1983, 2026

DURATION_TOL_S = 0.15    # source file vs phish.in track, same transfer
JOINT_CLIP_S = 2.5       # audio on each side of a joint in the rendered clips
MIN_MATCH_RATIO = 0.5    # tracks matched / tracks with audio to accept an item
FETCH_WORKERS = 8

# Taper labels for band members talking. Kept in step with the notes scanner
# that produced the first report (tmp/docs/find_missing_banter.rb).
BANTER_RE = re.compile(
    r"\b(banter|speaks?|talking|chatter|dialog(ue)?|monolog(ue)?|stage talk)\b"
    r"|\b(trey|fish|fishman|page|mike)\b.*\b(thanks|talks?|rap|story|speech|announcement)\b"
    r"|\b(talks?)\b.*\b(trey|fish|fishman|page|mike)\b",
    re.I)
NOT_BANTER_RE = re.compile(
    r"speak to me|no smoking|venue announcement|post show|thanks? (to|you)|thanx"
    r"|talking heads|conan", re.I)
# Lines that are only room noise, not anyone talking.
NOISE_ONLY_RE = re.compile(
    r"^[\s\W]*(set\s*\w*\s*)?(intro|outro|crowd|crowd noise|tuning|applause|encore break"
    r"|encore call|announcer|mic checks?|pa announcement)[\s\W]*$", re.I)
PROSE_RE = re.compile(r"[.!?]\s+\w|\b(was|were|is|are|had|has|the band|this|that|which)\b"
                      r".*\b(and|but|so|because)\b", re.I)
TRACK_TITLE_RE = re.compile(r"banter|speaks|talk|chatter|announcement|rap\b", re.I)


def load_junction_module():
    spec = importlib.util.spec_from_file_location("audio_junction_analysis", JUNCTION_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    sys.modules["audio_junction_analysis"] = module
    spec.loader.exec_module(module)
    return module


aja = load_junction_module()


# ---------------------------------------------------------------------------
# Notes scan (which shows to look at)


def banter_note_line(raw):
    line = raw.strip()
    if not line or len(line) > 70:
        return False
    if line.startswith(("*", "-", "#", "[", "note", "Note", "http", "Source", "Lineage")):
        return False
    if PROSE_RE.search(line) or NOISE_ONLY_RE.match(line) or NOT_BANTER_RE.search(line):
        return False
    if not BANTER_RE.search(line):
        return False
    words = len(re.sub(r"^[\W\d]+", "", line).split())
    return words <= 7 or bool(re.match(r"^\W*(d\d)?t?\d", line, re.I))


def note_hits(taper_notes):
    lines = [l.rstrip() for l in (taper_notes or "").splitlines()]
    hits = []
    for i, line in enumerate(lines):
        if not banter_note_line(line):
            continue
        prev = next((l.strip() for l in reversed(lines[:i]) if l.strip()), None)
        nxt = next((l.strip() for l in lines[i + 1:] if l.strip()), None)
        hits.append({"line": line.strip(), "prev": prev, "next": nxt})
    return hits


def notes_line_for(file_name, hits):
    """The notes entry for a source file, by its track token (d1t09 -> d1t09, t09, 09)."""
    stem = Path(file_name).stem.lower()
    m = re.search(r"(d\d+)?t(\d+)$", stem)
    if not m:
        return next((h["line"] for h in hits if stem in h["line"].lower()), None)
    disc, num = m.group(1), int(m.group(2))
    patterns = [rf"\b{disc}t0*{num}\b" if disc else None, rf"\bt0*{num}\b",
                rf"^\W*0*{num}[\.\)\-\s:]"]
    for pat in patterns:
        if not pat:
            continue
        for h in hits:
            if re.search(pat, h["line"], re.I):
                return h["line"]
    return None


def notes_line_by_neighbors(hits, after, before):
    """A notes entry sitting between lines that name the anchor tracks."""
    def names(track):
        if not track:
            return []
        return [re.sub(r"[^a-z0-9 ]", "", p.lower()).strip()
                for p in re.split(r"\s*-?>\s*", track["title"]) if p.strip()]
    def mentions(line, track):
        line = re.sub(r"[^a-z0-9 ]", "", (line or "").lower())
        return any(n and n in line for n in names(track))
    for h in hits:
        if (after and mentions(h["prev"], after)) or (before and mentions(h["next"], before)):
            return h["line"]
    return None


def fetch_show(date):
    resp = requests.get(f"{API_BASE}/shows/{date}", timeout=30)
    resp.raise_for_status()
    return resp.json()


def fetch_year_dates(year):
    dates = []
    page = 1
    while True:
        resp = requests.get(f"{API_BASE}/shows", timeout=60,
                            params={"year": year, "per_page": 100, "page": page})
        resp.raise_for_status()
        data = resp.json()
        dates += [s["date"] for s in data["shows"] if s.get("audio_status") != "missing"]
        if page >= data.get("total_pages", 1):
            return dates
        page += 1


def dates_from_notes():
    """Shows whose notes list more banter entries than they have banter tracks."""
    with ThreadPoolExecutor(FETCH_WORKERS) as pool:
        years = list(pool.map(fetch_year_dates, range(FIRST_YEAR, LAST_YEAR + 1)))
    all_dates = sorted(d for year in years for d in year)
    print(f"scanning taper notes of {len(all_dates)} shows", file=sys.stderr)

    def check(date):
        show = fetch_show(date)
        hits = note_hits(show.get("taper_notes"))
        if not hits:
            return None
        have = [t for t in show["tracks"]
                if TRACK_TITLE_RE.search(t["title"])
                or any(s["title"] == "Banter" for s in t.get("songs", []))]
        return date if len(hits) > len(have) else None

    with ThreadPoolExecutor(FETCH_WORKERS) as pool:
        flagged = [d for d in pool.map(check, all_dates) if d]
    print(f"{len(flagged)} shows flagged by notes", file=sys.stderr)
    return flagged


# ---------------------------------------------------------------------------
# archive.org source resolution


def parse_length(value):
    if value is None:
        return None
    s = str(value).strip()
    if re.fullmatch(r"\d+(\.\d+)?", s):
        return float(s)
    m = re.fullmatch(r"(?:(\d+):)?(\d+):(\d+(?:\.\d+)?)", s)
    if m:
        h, mi, se = m.groups()
        return int(h or 0) * 3600 + int(mi) * 60 + float(se)
    return None


def archive_items(date):
    resp = requests.get(f"{ARCHIVE}/advancedsearch.php", timeout=60, params={
        "q": f"phish AND date:[{date} TO {date}]",
        "fl[]": ["identifier", "title"], "rows": 50, "output": "json",
    })
    resp.raise_for_status()
    return [d["identifier"] for d in resp.json()["response"]["docs"]]


def archive_metadata(identifier):
    resp = requests.get(f"{ARCHIVE}/metadata/{identifier}", timeout=60)
    resp.raise_for_status()
    data = resp.json()
    files = []
    for f in data.get("files", []):
        name = f["name"]
        if not name.lower().endswith((".flac", ".shn", ".wav")):
            continue
        title = f.get("title") or Path(name).stem
        files.append({"name": name, "title": title, "length": parse_length(f.get("length"))})
    files.sort(key=lambda f: f["name"])
    checksums = [f["name"] for f in data.get("files", [])
                 if f["name"].lower().endswith((".ffp", ".md5", ".txt"))]
    return {"identifier": identifier, "files": files, "checksum_files": checksums,
            "description": data.get("metadata", {}).get("description", "")}


def match_item(files, tracks):
    """Map source files to phish.in tracks by duration. Returns
    (file_index -> track ids, track_id -> file indexes, matched track count)."""
    durations = {t["id"]: t["duration"] / 1000.0 for t in tracks if t.get("mp3_url")}
    file_to_tracks, track_to_files = {}, {}

    def close(a, b):
        return a is not None and abs(a - b) <= DURATION_TOL_S

    # One file, one track.
    for i, f in enumerate(files):
        for tid, dur in durations.items():
            if tid in track_to_files:
                continue
            if close(f["length"], dur):
                file_to_tracks[i] = [tid]
                track_to_files[tid] = [i]
                break
    # One track holding two or three consecutive files (TMWSIY > Avenu > TMWSIY).
    for tid, dur in durations.items():
        if tid in track_to_files:
            continue
        for i in range(len(files)):
            for n in (2, 3):
                group = files[i:i + n]
                if len(group) < n or any(j in file_to_tracks for j in range(i, i + n)):
                    continue
                lengths = [g["length"] for g in group]
                if None in lengths:
                    continue
                if close(sum(lengths), dur):
                    for j in range(i, i + n):
                        file_to_tracks[j] = [tid]
                    track_to_files[tid] = list(range(i, i + n))
                    break
            if tid in track_to_files:
                break
    # One file split into two or three consecutive tracks.
    ordered = [t for t in sorted(tracks, key=lambda t: t["position"]) if t["id"] in durations]
    for i, f in enumerate(files):
        if i in file_to_tracks or f["length"] is None:
            continue
        for k in range(len(ordered)):
            for n in (2, 3):
                group = ordered[k:k + n]
                if len(group) < n or any(t["id"] in track_to_files for t in group):
                    continue
                if close(sum(durations[t["id"]] for t in group), f["length"]):
                    file_to_tracks[i] = [t["id"] for t in group]
                    for t in group:
                        track_to_files[t["id"]] = [i]
                    break
            if i in file_to_tracks:
                break
    return file_to_tracks, track_to_files, len(track_to_files), len(durations)


def local_item(date, out_dir):
    """A hand-fetched source (e.g. from the Phish Spreadsheet) dropped into
    <out-dir>/<date>/source/, presented like an archive.org item."""
    src_dir = out_dir / date / "source"
    if not src_dir.is_dir():
        return None
    files = []
    for path in sorted(src_dir.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in (".flac", ".shn", ".wav", ".mp3"):
            continue
        try:
            length = aja.probe_duration(path)
        except Exception:
            length = None
        out = aja.run(["ffprobe", "-v", "error", "-show_entries", "format_tags=title",
                       "-of", "default=noprint_wrappers=1:nokey=1", str(path)])
        title = out.stdout.decode().strip() or path.stem
        files.append({"name": str(path.relative_to(src_dir)), "title": title,
                      "length": length, "path": str(path)})
    if not files:
        return None
    return {"identifier": f"local:{src_dir}", "files": files, "checksum_files": [],
            "description": "", "local": True}


def resolve_source(date, tracks, out_dir):
    """Best source for the show - a local drop-in, else the archive.org item
    whose file lengths match phish.in - or None with the candidates tried."""
    tried = []
    best = None
    local = local_item(date, out_dir)
    if local:
        f2t, t2f, matched, total = match_item(local["files"], tracks)
        tried.append({"identifier": local["identifier"], "matched": matched, "total": total,
                      "files": len(local["files"])})
        if total and matched / total >= MIN_MATCH_RATIO:
            return {"meta": local, "file_to_tracks": f2t, "track_to_files": t2f,
                    "matched": matched, "total": total}, tried
    for identifier in archive_items(date):
        try:
            meta = archive_metadata(identifier)
        except requests.RequestException as e:
            tried.append({"identifier": identifier, "error": str(e)})
            continue
        f2t, t2f, matched, total = match_item(meta["files"], tracks)
        ratio = matched / total if total else 0.0
        tried.append({"identifier": identifier, "matched": matched, "total": total,
                      "files": len(meta["files"])})
        if ratio >= MIN_MATCH_RATIO and (best is None or matched > best["matched"]):
            best = {"meta": meta, "file_to_tracks": f2t, "track_to_files": t2f,
                    "matched": matched, "total": total}
    return best, tried


def spreadsheet_row(date, out_dir):
    """Download link, lineage and notes for a date from the Phish Spreadsheet."""
    import datetime
    import openpyxl
    path = out_dir / "phish_spreadsheet.xlsx"
    if not path.exists():
        resp = requests.get(f"{SPREADSHEET_URL}/export?format=xlsx", timeout=120)
        resp.raise_for_status()
        path.write_bytes(resp.content)
    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)
    year = date[:4]
    sheet = wb[year] if year in wb.sheetnames else wb["'83-87"] if "'83-87" in wb.sheetnames else None
    if sheet is None:
        return None
    want = datetime.date.fromisoformat(date)
    for row in sheet.iter_rows(values_only=True):
        first = row[0]
        if isinstance(first, datetime.datetime):
            first = first.date()
        if first != want:
            continue
        cells = [c for c in row if c is not None]
        link = next((m.group(1) for c in cells
                     for m in [re.search(r'HYPERLINK\("([^"]+)"', str(c))]
                     if m and "phish.net" not in m.group(1)), None)
        return {"link": link, "source": str(row[8]) if len(row) > 8 and row[8] else "",
                "notes": str(row[9]) if len(row) > 9 and row[9] else ""}
    return None


# ---------------------------------------------------------------------------
# Download + verify


def flac_header_md5(path):
    with open(path, "rb") as fh:
        head = fh.read(42)
    if head[:4] != b"fLaC" or (head[4] & 0x7F) != 0:
        return None
    return head[26:42].hex()


def expected_md5s(identifier, checksum_files):
    """{"ffp": {name: audio md5}, "file": {name: whole-file md5}} from the item's
    checksum listings. An .ffp fingerprints the decoded audio (what the FLAC
    STREAMINFO header carries); an .md5 hashes the file bytes."""
    found = {"ffp": {}, "file": {}}
    for name in checksum_files:
        try:
            text = requests.get(f"{ARCHIVE}/download/{identifier}/{name}", timeout=60).text
        except requests.RequestException:
            continue
        for m in re.finditer(r"([^\s:*]+\.flac)\s*:\s*([0-9a-fA-F]{32})", text):
            found["ffp"][Path(m.group(1)).name] = m.group(2).lower()
        for m in re.finditer(r"([0-9a-fA-F]{32})\s+\*?([^\s]+\.flac)", text):
            found["file"][Path(m.group(2)).name] = m.group(1).lower()
    return found


def verify_download(path, md5s):
    """True/False against the item's checksums, None when it lists none."""
    name = Path(path).name
    if name in md5s["ffp"]:
        return flac_header_md5(path) == md5s["ffp"][name]
    if name in md5s["file"]:
        h = hashlib.md5()
        with open(path, "rb") as fh:
            for chunk in iter(lambda: fh.read(1 << 20), b""):
                h.update(chunk)
        return h.hexdigest() == md5s["file"][name]
    return None


def download_source_file(identifier, name, dest_dir):
    dest = dest_dir / Path(name).name
    if dest.exists():
        return dest
    dest_dir.mkdir(parents=True, exist_ok=True)
    url = f"{ARCHIVE}/download/{identifier}/{name}"
    print(f"  downloading {url}", file=sys.stderr)
    tmp = dest.with_suffix(dest.suffix + ".part")
    # archive.org's mirrors throw intermittent 5xx; a retry usually lands elsewhere.
    for attempt in range(3):
        try:
            with requests.get(url, stream=True, timeout=300) as resp:
                resp.raise_for_status()
                with open(tmp, "wb") as fh:
                    for chunk in resp.iter_content(1 << 20):
                        fh.write(chunk)
            tmp.rename(dest)
            return dest
        except requests.RequestException as e:
            if attempt == 2:
                raise
            print(f"  retrying after {e}", file=sys.stderr)
    return dest


# ---------------------------------------------------------------------------
# Joints


@dataclass
class Edge:
    head: np.ndarray
    tail: np.ndarray


def load_track_edges(mp3_url):
    path = aja.download_audio(mp3_url)
    duration = aja.probe_duration(path)
    window = min(aja.EDGE_WINDOW_S, duration)
    return Edge(head=aja.decode_segment(path, 0, window),
                tail=aja.decode_segment(path, max(duration - window, 0), window))


def load_file_edges(path):
    duration = aja.probe_duration(path)
    window = min(aja.EDGE_WINDOW_S, duration)
    return Edge(head=aja.decode_segment(path, 0, window),
                tail=aja.decode_segment(path, max(duration - window, 0), window))


def render_preview(src, out_path):
    """Listenable mp3 of the whole source file (128k; the FLAC stays on disk)."""
    if not out_path.exists():
        aja.run(["ffmpeg", "-v", "error", "-y", "-i", str(src), "-b:a", "128k", str(out_path)])
    return out_path.name


def joint(tail_edge, head_edge, clip_path):
    j = aja.score_junction(tail_edge.tail, head_edge.head)
    n = int(JOINT_CLIP_S * aja.SAMPLE_RATE)
    spliced = np.concatenate([tail_edge.tail[-n:], head_edge.head[:n]]).astype(np.float32)
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".f32") as tmp:
        tmp.write(spliced.tobytes())
        tmp.flush()
        aja.run(["ffmpeg", "-v", "error", "-y", "-f", "f32le", "-ar", str(aja.SAMPLE_RATE),
                 "-ac", "1", "-i", tmp.name, "-b:a", "128k", str(clip_path)])
    return {"score": round(j.score, 2), "verdict": aja.classify(j.score),
            "detail": j.parts(), "clip": clip_path.name}


# ---------------------------------------------------------------------------
# Per-show analysis


@dataclass
class Candidate:
    date: str
    key: str
    source_item: str
    source_file: str
    source_title: str
    length_s: float
    source_path: str
    md5_verified: bool | None
    after_id: int | None
    after_title: str | None
    after_position: int | None
    before_id: int | None
    before_title: str | None
    before_position: int | None
    status: str                      # candidate | already_present | unanchored
    note: str = ""
    joints: dict = field(default_factory=dict)
    notes_line: str | None = None
    after_set: str | None = None
    before_set: str | None = None
    preview: str | None = None       # rendered mp3 of the whole source file, in clips/
    members: list = field(default_factory=list)  # every file in the run (1 for a lone file)


@dataclass
class ShowReport:
    date: str
    venue: str
    url: str
    note_hits: list
    source_item: str | None
    matched: int
    total: int
    tried: list
    unmatched_files: list
    candidates: list
    error: str | None = None
    spreadsheet: dict | None = None
    filler_skipped: list = field(default_factory=list)


def analyze_show(date, out_dir):
    show = fetch_show(date)
    tracks = sorted(show["tracks"], key=lambda t: t["position"])
    by_id = {t["id"]: t for t in tracks}
    hits = note_hits(show.get("taper_notes"))
    report = ShowReport(date=date, venue=show.get("venue_name") or "",
                        url=f"{SITE_BASE}/{date}", note_hits=hits, source_item=None,
                        matched=0, total=0, tried=[], unmatched_files=[], candidates=[])
    best, tried = resolve_source(date, tracks, out_dir)
    report.tried = tried
    if best is None:
        report.error = ("no archive.org item for this date" if not tried
                        else "no archive.org item matches phish.in's track durations")
        try:
            report.spreadsheet = spreadsheet_row(date, out_dir)
        except Exception as e:
            report.spreadsheet = {"link": None, "source": "", "notes": f"lookup failed: {e}"}
        return report

    meta = best["meta"]
    files = meta["files"]
    f2t = best["file_to_tracks"]
    report.source_item = meta["identifier"]
    report.matched, report.total = best["matched"], best["total"]
    # Filler is bonus material tacked onto the end of an item (a demo, another
    # show), not part of this performance. When the item or the notes call it
    # out, the trailing run of unmatched files is it.
    filler_text = f"{meta.get('description', '')}\n{show.get('taper_notes') or ''}"
    report.filler_skipped = []
    if re.search(r"\b(filler|bonus)\b", filler_text, re.I) and f2t:
        last_matched = max(f2t)
        trailing = [i for i in range(last_matched + 1, len(files))]
        report.filler_skipped = [files[i]["title"] for i in trailing]
        files = files[:last_matched + 1]
    report.unmatched_files = [
        {"name": f["name"], "title": f["title"], "length": f["length"],
         "banter": bool(BANTER_RE.search(f["title"]) and not NOT_BANTER_RE.search(f["title"]))}
        for i, f in enumerate(files) if i not in f2t
    ]

    show_dir = out_dir / date
    clip_dir = out_dir / "clips"
    clip_dir.mkdir(parents=True, exist_ok=True)
    md5s = None
    edge_cache = {}

    def track_edges(tid):
        if tid not in edge_cache:
            edge_cache[tid] = load_track_edges(by_id[tid]["mp3_url"])
        return edge_cache[tid]

    def fetch(f):
        return (Path(f["path"]) if meta.get("local")
                else download_source_file(meta["identifier"], f["name"], show_dir))

    def checksums():
        nonlocal md5s
        if md5s is None:
            md5s = (expected_md5s(meta["identifier"], meta["checksum_files"])
                    if not meta.get("local") else {"ffp": {}, "file": {}})
        return md5s

    # Consecutive unmatched files form one run: a lone banter file, or a whole
    # stretch (even a set) phish.in never got. Joints are only meaningful at a
    # run's ends, so only its first and last files are fetched.
    runs, i = [], 0
    while i < len(files):
        if i in f2t:
            i += 1
            continue
        j = i
        while j + 1 < len(files) and (j + 1) not in f2t:
            j += 1
        runs.append((i, j))
        i = j + 1

    for first, last in runs:
        prev_idx = next((j for j in range(first - 1, -1, -1) if j in f2t), None)
        next_idx = next((j for j in range(last + 1, len(files)) if j in f2t), None)
        after = by_id[f2t[prev_idx][-1]] if prev_idx is not None else None
        before = by_id[f2t[next_idx][0]] if next_idx is not None else None
        members = files[first:last + 1]
        f = members[0]
        key = f"{date}:{Path(f['name']).stem}"
        cand = Candidate(
            date=date, key=key, source_item=meta["identifier"], source_file=f["name"],
            source_title=f["title"], length_s=f["length"] or 0.0, source_path="",
            md5_verified=None,
            after_id=after and after["id"], after_title=after and after["title"],
            after_position=after and after["position"],
            before_id=before and before["id"], before_title=before and before["title"],
            before_position=before and before["position"],
            after_set=after and after.get("set_name"),
            before_set=before and before.get("set_name"),
            status="candidate" if len(members) == 1 else "run",
            members=[{"name": m["name"], "title": m["title"], "length": m["length"]}
                     for m in members],
        )
        cand.notes_line = (notes_line_for(f["name"], hits)
                           or notes_line_by_neighbors(hits, after, before))
        if after is None or before is None:
            if cand.status == "candidate":
                cand.status = "unanchored"
        elif before["position"] != after["position"] + 1:
            between = [t for t in tracks if after["position"] < t["position"] < before["position"]]
            same = [t for t in between
                    if abs(t["duration"] / 1000.0 - (f["length"] or -1)) <= DURATION_TOL_S]
            if same and len(members) == 1:
                cand.status = "already_present"
                cand.note = f"phish.in has {same[0]['title']} ({same[0]['duration'] / 1000:.1f}s) here"
            else:
                cand.note = "phish.in already has " + \
                    ", ".join(t["title"] for t in between) + " between these"
        if cand.status == "already_present":
            report.candidates.append(cand)
            continue

        head_path = fetch(members[0])
        tail_path = head_path if len(members) == 1 else fetch(members[-1])
        cand.source_path = str(head_path)
        cand.md5_verified = verify_download(head_path, checksums())
        stem = re.sub(r"[^a-z0-9]+", "-", key.lower())
        cand.preview = render_preview(head_path, clip_dir / f"{stem}-full.mp3")
        if len(members) > 1:
            cand.members[-1]["preview"] = render_preview(
                tail_path, clip_dir / f"{stem}-last-full.mp3")
        head_edge = load_file_edges(head_path)
        tail_edge = head_edge if len(members) == 1 else load_file_edges(tail_path)
        cand.joints = {}
        if after:
            cand.joints["in"] = joint(track_edges(after["id"]), head_edge, clip_dir / f"{stem}-in.mp3")
        if before:
            cand.joints["out"] = joint(tail_edge, track_edges(before["id"]), clip_dir / f"{stem}-out.mp3")
        if after and before:
            cand.joints["direct"] = joint(track_edges(after["id"]), track_edges(before["id"]),
                                          clip_dir / f"{stem}-direct.mp3")
        report.candidates.append(cand)
    return report


# ---------------------------------------------------------------------------
# Review page


def esc(s):
    return html.escape(str(s if s is not None else ""))


def embedded_fonts():
    font_dir = REPO_ROOT / "node_modules/@fontsource/open-sans-condensed/files"
    faces = []
    for weight in (300, 700):
        path = font_dir / f"open-sans-condensed-latin-{weight}-normal.woff2"
        if not path.exists():
            continue
        b64 = base64.b64encode(path.read_bytes()).decode()
        faces.append(
            '@font-face { font-family: "Open Sans Condensed"; font-style: normal; '
            f'font-weight: {weight}; font-display: swap; '
            f'src: url(data:font/woff2;base64,{b64}) format("woff2"); ' + "}")
    return "\n  ".join(faces)


def tried_summary(t):
    if t.get("error"):
        return t["error"]
    return f"{t['matched']}/{t['total']} tracks matched, {t['files']} files"


def fmt_len(seconds):
    seconds = int(round(seconds or 0))
    return f"{seconds // 60}:{seconds % 60:02d}"


def player(name, cls="full"):
    if not name:
        return '<span class="noaudio">no audio</span>'
    return f'<audio class="{cls}" controls preload="none" src="clips/{esc(name)}"></audio>'


def joint_block(label, j):
    if not j:
        return f'<div class="joint none"><span class="jlabel">{label}</span></div>'
    return (f'<div class="joint {j["verdict"].lower()}" title="{esc(j["detail"])}">'
            f'<span class="jlabel">{label}</span>{player(j["clip"], "jaudio")}</div>')


_slugs = {}


def track_link(date, track_id, title):
    """Title linked to its phish.in page; slugs come from the show API once per date."""
    if date not in _slugs:
        try:
            _slugs[date] = {t["id"]: t.get("slug") for t in fetch_show(date)["tracks"]}
        except requests.RequestException:
            _slugs[date] = {}
    slug = _slugs[date].get(track_id)
    if not slug:
        return esc(title)
    return f'<a href="{SITE_BASE}/{date}/{slug}" target="_blank">{esc(title)}</a>'


def placement_text(c):
    if c.after_id is None and c.before_id is not None:
        return f"Start of {esc(c.before_set)}"
    if c.before_id is None and c.after_id is not None:
        return f"End of {esc(c.after_set)}"
    return (f'<span class="k">Before</span> {track_link(c.date, c.after_id, c.after_title)}'
            f' <span class="k">After</span> {track_link(c.date, c.before_id, c.before_title)}')


def source_block(c):
    if len(c.members) > 1:
        first, last = c.members[0], c.members[-1]
        middle = ", ".join(esc(m["title"]) for m in c.members[1:-1])
        return (f'<div class="src"><div class="srcname">{esc(first["title"])}'
                f' <span class="meta">{fmt_len(first["length"])}</span></div>{player(c.preview)}'
                f'{"<div class=middle>" + middle + "</div>" if middle else ""}'
                f'<div class="srcname">{esc(last["title"])}'
                f' <span class="meta">{fmt_len(last["length"])}</span></div>'
                f'{player(last.get("preview"))}</div>')
    return f'<div class="src">{player(c.preview)}</div>'


def candidate_row(c):
    md5 = {True: "md5 ok", False: "MD5 MISMATCH", None: "no checksum"}[c.md5_verified]
    title = (f"{len(c.members)} consecutive files not on phish.in" if len(c.members) > 1
             else c.source_title)
    chips = [f'<span class="chip place">{placement_text(c)}</span>']
    if c.notes_line:
        chips.append(f'<span class="chip notes" title="taper notes">{esc(c.notes_line)}</span>')
    if c.note:
        chips.append(f'<span class="chip warn">{esc(c.note)}</span>')
    meta = [esc(c.source_file)]
    if c.status in ("candidate", "unanchored", "run"):
        meta.append(md5)
    head = (f'<div class="head"><strong>{esc(title)}</strong>{"".join(chips)}'
            f'<span class="dur">{fmt_len(c.length_s) if len(c.members) == 1 else ""}</span></div>')
    if c.status == "already_present":
        return (f'<div class="row already"><div class="head"><strong>{esc(title)}</strong>'
                f'<span class="chip">already on phish.in</span>'
                f'<span class="meta">{esc(c.note)}</span></div></div>')
    joints = (f'<div class="joints">{joint_block("Joint in", c.joints.get("in"))}'
              f'{joint_block("Joint out", c.joints.get("out"))}'
              f'{joint_block("Direct", c.joints.get("direct"))}</div>')
    decide = ""
    payload = ""
    if c.status in ("candidate", "unanchored"):
        payload = ' data-payload="' + esc(json.dumps({
            "date": c.date, "key": c.key, "after_id": c.after_id, "after_title": c.after_title,
            "after_position": c.after_position, "before_id": c.before_id,
            "before_title": c.before_title, "before_position": c.before_position,
            "source_path": c.source_path, "source_item": c.source_item,
            "source_file": c.source_file,
        })) + '"'
        decide = ('<div class="decide"><label><input type="checkbox" class="approve"> Approve</label>'
                  '<label><input type="checkbox" class="skip"> Skip</label>'
                  '<label class="field"><span>Title</span><input class="title" value="Banter"></label>'
                  '<label class="field"><span>Song</span><select class="song"></select></label></div>')
    cls = "row " + ("cand" if c.status in ("candidate", "unanchored") else c.status)
    return (f'<div class="{cls}"{payload}>{head}<div class="body">{source_block(c)}'
            f'<div class="meta small">{" · ".join(meta)}</div>{joints}</div>{decide}</div>')


def fetch_songs():
    try:
        resp = requests.get(f"{API_BASE}/songs", timeout=60,
                            params={"per_page": 2000, "sort": "title:asc"})
        resp.raise_for_status()
        songs = [(sg["id"], sg["title"]) for sg in resp.json()["songs"]]
    except requests.RequestException as e:
        print(f"song list unavailable ({e}); dropdown will only offer Banter", file=sys.stderr)
        songs = []
    if not any(t == "Banter" for _, t in songs):
        songs.append((886, "Banter"))
    return songs


def write_review(out_dir, reports):
    reports = sorted(reports, key=lambda r: r.date)
    songs = fetch_songs()
    song_options = "".join(
        f'<option value="{sid}"{" selected" if title == "Banter" else ""}>{esc(title)}</option>'
        for sid, title in songs)
    sections = []
    n_cand = 0
    for r in reports:
        hits = "".join(f'<li><code>{esc(h["line"])}</code> <span class="meta">after '
                       f'<code>{esc(h["prev"])}</code>, before <code>{esc(h["next"])}</code></span></li>'
                       for h in r.note_hits)
        notes = f'<details class="notes"><summary>Taper notes entries ({len(r.note_hits)})</summary><ul>{hits}</ul></details>' if r.note_hits else ""
        head = (f'<div class="setsep" id="{r.date}"><span class="setsep-name">{r.date}</span>'
                f'<span class="venue">{esc(r.venue)}</span>'
                f'<a class="ext" href="{r.url}" target="_blank">phish.in</a>'
                f'<a class="ext" href="https://phish.net/setlists/?d={r.date}" target="_blank">phish.net</a>')
        if r.error:
            tried = "".join(
                f"<li><a href='{ARCHIVE}/details/{esc(t['identifier'])}' target='_blank'>"
                f"{esc(t['identifier'])}</a>: {esc(tried_summary(t))}</li>" for t in r.tried)
            sheet = ""
            if r.spreadsheet:
                link = (f"<a href='{esc(r.spreadsheet['link'])}' target='_blank'>download</a>"
                        if r.spreadsheet.get("link") else "no download link")
                sheet = (f"<p>Phish Spreadsheet: {link}; source <code>{esc(r.spreadsheet.get('source'))}"
                         f"</code> {esc(r.spreadsheet.get('notes'))}. Extract it into "
                         f"<code>data/banter_scan/{r.date}/source/</code> and re-run "
                         f"<code>rake banter_scan:run[{r.date}]</code>.</p>")
            elif not [t for t in r.tried if not t["identifier"].startswith("local:")]:
                sheet = f"<p>Not in the Phish Spreadsheet either (<a href='{SPREADSHEET_URL}'>sheet</a>).</p>"
            sections.append(f'<section class="unresolved">{head}<span class="chip warn">unresolved</span></div>'
                            f"{notes}<p class='meta'>{esc(r.error)}.</p>"
                            f"<ul class='meta'>{tried}</ul>{sheet}</section>")
            continue
        src_link = (esc(r.source_item) if r.source_item.startswith("local:") else
                    f"<a href='{ARCHIVE}/details/{esc(r.source_item)}' target='_blank'>{esc(r.source_item)}</a>")
        src = (f'<div class="source">Source {src_link} '
               f'<span class="chip">{r.matched}/{r.total} tracks matched</span>')
        if r.filler_skipped:
            src += (f' <span class="chip" title="{esc(", ".join(r.filler_skipped))}">'
                    f'{len(r.filler_skipped)} trailing filler files skipped</span>')
        src += "</div>"
        rows = [candidate_row(c) for c in r.candidates]
        n_cand += sum(1 for c in r.candidates if c.status in ("candidate", "unanchored"))
        body = "".join(rows) or '<p class="meta">Every source file matches a phish.in track.</p>'
        sections.append(f"<section>{head}</div>{src}{notes}{body}</section>")

    (out_dir / "review.html").write_text(f"""<!doctype html>
<meta charset="utf-8">
<title>Banter placement review</title>
<style>
  {embedded_fonts()}
  :root {{
    --bg: #ffffff; --header: #f4f4f6; --fg: #1c1c1e; --muted: #6b6b70;
    --line: #e3e3e7; --card: #fafafa; --sel: #eef2f8; --sel-line: #b9cbe6;
    --accent: #b8860b; --ok: #1a6b2f; --warn: #b8860b; --err: #b00020;
    --ok-bg: #e6f4ea; --warn-bg: #fdf3d9; --err-bg: #fbe4e6;
    --btn: #f2f2f4; --btn-line: #d8d8dd; --link: #2f6fd0; --btn-hover: #e8e8ec;
  }}
  @media (prefers-color-scheme: dark) {{
    :root {{
      --bg: #1b1f27; --header: #12151b; --fg: #e6e9ef; --muted: #8b95a7;
      --line: #2f3542; --card: #232833; --sel: #26313f; --sel-line: #3a5578;
      --accent: #e0a92e; --ok: #5fce85; --warn: #e0a92e; --err: #ff6b81;
      --ok-bg: #1f3a2a; --warn-bg: #3d3320; --err-bg: #452229;
      --btn: #2e3644; --btn-line: #4a5568; --link: #7fb3f0; --btn-hover: #3a4354;
    }}
    audio {{ color-scheme: dark; }}
  }}
  * {{ box-sizing: border-box; }}
  body {{ font: 14px/1.5 -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif;
          margin: 0 auto 3rem; max-width: 1300px; padding: 6.2rem 1.5rem 0;
          background: var(--bg); color: var(--fg); -webkit-font-smoothing: antialiased; }}
  h1 {{ font-family: "Open Sans Condensed", -apple-system, sans-serif; font-weight: 700;
        letter-spacing: .01em; font-size: 21px; margin: 0; flex: 1 1 auto; }}
  h1 .count {{ color: var(--muted); font-weight: 400; font-size: 15px; }}
  a {{ color: var(--link); text-decoration-color: color-mix(in srgb, var(--link) 45%, transparent);
       text-underline-offset: 2px; }}
  a:hover {{ text-decoration-color: currentColor; }}
  code {{ font-size: 12px; }}
  .meta {{ color: var(--muted); }}
  .small {{ font-size: 12px; }}
  #topbar {{ position: fixed; top: 0; left: 0; right: 0; z-index: 10; background: var(--header);
             border-bottom: 1px solid var(--line); padding: .55rem 1.5rem .35rem; }}
  #topbar .row1 {{ display: flex; align-items: center; gap: 1rem; max-width: 1300px; margin: 0 auto; }}
  #legend {{ display: flex; align-items: baseline; gap: .9rem; font-size: 13px;
             font-variant-numeric: tabular-nums; }}
  #legend .pct {{ font-size: 20px; font-weight: 700;
                  font-family: "Open Sans Condensed", -apple-system, sans-serif; }}
  #legend .k {{ color: var(--muted); margin-right: .3rem; }}
  #legend .v {{ font-weight: 600; }}
  #topbar .bar {{ height: 4px; border-radius: 2px; background: var(--line); margin: .45rem auto 0;
                  max-width: 1300px; overflow: hidden; font-size: 0; }}
  #topbar .bar i {{ display: inline-block; height: 100%; vertical-align: top; }}
  #topbar .bar .fill-ok {{ background: var(--ok); }}
  #topbar .bar .fill-skip {{ background: var(--accent); opacity: .55; }}
  #export {{ padding: .4rem 1rem; font: inherit; font-weight: 600; cursor: pointer; background: var(--btn);
             color: var(--fg); border: 1px solid var(--btn-line); border-radius: 7px; }}
  #export:hover {{ background: var(--btn-hover); border-color: var(--muted); }}
  .key {{ display: flex; gap: 1rem; font-size: 12px; color: var(--muted); margin-bottom: 1.5rem; }}
  .key i {{ display: inline-block; width: 10px; height: 10px; border-radius: 3px; margin-right: .3rem;
            vertical-align: -1px; }}
  .setsep {{ display: flex; align-items: center; gap: .6rem; margin: 2.2rem 0 .3rem;
             padding: 0 1rem .35rem; border-bottom: 2px solid var(--line); }}
  .setsep-name {{ font-family: "Open Sans Condensed", -apple-system, sans-serif; font-size: 22px;
                  font-weight: 700; letter-spacing: .01em; font-variant-numeric: tabular-nums; }}
  .venue {{ color: var(--muted); font-size: 14px; }}
  .ext {{ font-size: 12px; text-decoration: none; border: 1px solid var(--line); border-radius: 999px;
          padding: .05rem .55rem; }}
  .ext:hover {{ border-color: var(--link); }}
  .source {{ color: var(--muted); font-size: 13px; padding: .35rem 1rem 0; }}
  details.notes {{ padding: .2rem 1rem; font-size: 13px; color: var(--muted); }}
  details.notes ul {{ margin: .3rem 0 .5rem; }}
  section.unresolved p, section.unresolved ul {{ padding: 0 1rem; margin: .3rem 0; font-size: 13px; }}
  .chip {{ color: var(--muted); font-size: 12px; border: 1px solid var(--line); border-radius: 999px;
           padding: .05rem .55rem; white-space: nowrap; max-width: 34rem; overflow: hidden;
           text-overflow: ellipsis; }}
  .chip.place {{ color: var(--fg); }}
  .chip.place a {{ color: var(--fg); text-decoration-color: color-mix(in srgb, var(--fg) 30%, transparent); }}
  .chip.place a:hover {{ color: var(--link); }}
  .chip .k {{ color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: .04em;
              margin: 0 .15rem 0 .35rem; }}
  .chip .k:first-child {{ margin-left: 0; }}
  .chip.warn {{ color: var(--warn); border-color: color-mix(in srgb, var(--warn) 50%, transparent); }}
  .row {{ border: 1px solid transparent; border-bottom-color: var(--line); border-radius: 10px;
          padding: .85rem 1rem; transition: background .12s ease; }}
  .row:hover {{ background: var(--card); }}
  .row.sel {{ background: var(--sel); border-color: var(--sel-line);
              box-shadow: inset 3px 0 0 var(--link); }}
  .row.done:not(.sel) .body, .row.done:not(.sel) .decide .field {{ display: none; }}
  .row.done:not(.sel) .head {{ margin-bottom: 0; }}
  .row.done:not(.sel) .decide {{ margin-top: .2rem; }}
  .row.skipped:not(.sel) {{ opacity: .55; }}
  .row.skipped:not(.sel) .head strong {{ text-decoration: line-through; }}
  .row.approved:not(.sel) .head strong {{ color: var(--ok); }}
  .head {{ cursor: pointer; }}
  .skip {{ accent-color: var(--muted); }}
  kbd {{ font: 11px/1 -apple-system, sans-serif; border: 1px solid var(--line); border-radius: 4px;
         padding: .1rem .3rem; color: var(--muted); }}
  .row.already {{ opacity: .55; }}
  .row.run {{ opacity: .85; }}
  .head {{ display: flex; gap: .6rem; align-items: center; flex-wrap: wrap; margin-bottom: .5rem; }}
  .head strong {{ font-family: "Open Sans Condensed", -apple-system, sans-serif; font-size: 21px;
                  font-weight: 700; }}
  .dur {{ color: var(--muted); font-variant-numeric: tabular-nums; font-size: 13px; margin-left: auto; }}
  .src {{ margin: .2rem 0; }}
  .srcname {{ font-size: 13px; margin-top: .4rem; }}
  .middle {{ color: var(--muted); font-size: 12px; margin: .3rem 0; }}
  audio.full {{ width: 100%; height: 32px; display: block; }}
  .noaudio {{ color: var(--muted); font-size: 12px; }}
  .joints {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: .6rem; margin-top: .6rem; }}
  .joint {{ border-radius: 8px; padding: .45rem .6rem .5rem; background: var(--card);
            border-left: 4px solid var(--line); }}
  .joint.continuous {{ background: var(--ok-bg); border-left-color: var(--ok); }}
  .joint.suspect {{ background: var(--warn-bg); border-left-color: var(--warn); }}
  .joint.broken {{ background: var(--err-bg); border-left-color: var(--err); }}
  .joint.none {{ opacity: .5; }}
  .jlabel {{ display: block; font-family: "Open Sans Condensed", -apple-system, sans-serif;
             font-size: 13px; font-weight: 700; text-transform: uppercase; letter-spacing: .05em;
             color: var(--muted); margin-bottom: .25rem; }}
  .joint.continuous .jlabel {{ color: var(--ok); }}
  .joint.suspect .jlabel {{ color: var(--warn); }}
  .joint.broken .jlabel {{ color: var(--err); }}
  audio.jaudio {{ width: 100%; height: 32px; display: block; }}
  .decide {{ display: flex; align-items: center; gap: .8rem; margin-top: .7rem; font-size: 13px; }}
  .decide label {{ display: flex; align-items: center; gap: .4rem; font-weight: 600; cursor: pointer; }}
  input[type="checkbox"] {{ width: 1.15rem; height: 1.15rem; cursor: pointer; accent-color: var(--link); }}
  .decide input[type="text"], .decide input:not([type]) {{ font: inherit; font-size: 13px; padding: .3rem .5rem;
      background: var(--bg); color: var(--fg); border: 1px solid var(--btn-line); border-radius: 6px; }}
  .decide .field {{ font-weight: 400; gap: .4rem; }}
  .decide .field span {{ color: var(--muted); font-size: 11px; text-transform: uppercase;
                         letter-spacing: .04em; }}
  .decide .title {{ width: 14rem; }}
  .decide select {{ font: inherit; font-size: 13px; padding: .3rem .4rem; max-width: 18rem;
                    background: var(--bg); color: var(--fg); border: 1px solid var(--btn-line);
                    border-radius: 6px; }}
</style>
<div id="topbar"><div class="row1">
  <h1>Banter placement review <span class="count">{len(reports)} shows · {n_cand} candidates</span></h1>
  <div id="legend"><span class="pct">0%</span>
    <span><span class="k">approved</span><span class="v" id="n-approved">0</span></span>
    <span><span class="k">skipped</span><span class="v" id="n-skipped">0</span></span>
    <span><span class="k">of</span><span class="v">{n_cand}</span></span></div>
  <button id="export">Export approved.json</button>
</div><div class="bar"><i class="fill-ok" style="width:0"></i><i class="fill-skip" style="width:0"></i></div></div>
<template id="songs">{song_options}</template>
<div class="key"><span><i style="background:var(--ok)"></i>continuous</span>
  <span><i style="background:var(--warn)"></i>suspect</span>
  <span><i style="background:var(--err)"></i>broken</span>
  <span>Joint in: track before &gt; file. Joint out: file &gt; track after. Direct: what phish.in has now.</span>
  <span><kbd>&larr;</kbd><kbd>&rarr;</kbd> 1s on the last player (<kbd>shift</kbd> 5s) · <kbd>space</kbd> play/pause ·
  <kbd>w</kbd> approve · <kbd>k</kbd> skip · <kbd>&uarr;</kbd><kbd>&darr;</kbd> rows</span></div>
{''.join(sections)}
<script>
const KEY = "banter_scan_progress";
const rows = () => [...document.querySelectorAll(".row.cand")];
const payloadOf = r => JSON.parse(r.dataset.payload);
const songOptions = document.getElementById("songs").content;
rows().forEach(r => r.querySelector("select.song").appendChild(songOptions.cloneNode(true)));
const isApproved = r => r.querySelector("input.approve").checked;
const isSkipped = r => r.querySelector("input.skip").checked;
function save() {{
  const state = {{}};
  rows().forEach(r => {{
    state[payloadOf(r).key] = {{
      approved: isApproved(r), skipped: isSkipped(r),
      title: r.querySelector("input.title").value,
      song_id: r.querySelector("select.song").value,
    }};
  }});
  try {{ localStorage.setItem(KEY, JSON.stringify(state)); }} catch (e) {{}}
}}
function restore() {{
  let state = {{}};
  try {{ state = JSON.parse(localStorage.getItem(KEY) || "{{}}"); }} catch (e) {{}}
  rows().forEach(r => {{
    const s = state[payloadOf(r).key];
    if (!s) return;
    r.querySelector("input.approve").checked = !!s.approved;
    r.querySelector("input.skip").checked = !!s.skipped;
    if (s.title) r.querySelector("input.title").value = s.title;
    if (s.song_id) r.querySelector("select.song").value = s.song_id;
  }});
}}
function refresh() {{
  const all = rows();
  all.forEach(r => {{
    r.classList.toggle("approved", isApproved(r));
    r.classList.toggle("skipped", isSkipped(r));
    r.classList.toggle("done", isApproved(r) || isSkipped(r));
  }});
  const approved = all.filter(isApproved).length, skipped = all.filter(isSkipped).length;
  document.getElementById("n-approved").textContent = approved;
  document.getElementById("n-skipped").textContent = skipped;
  const pct = all.length ? Math.round((approved + skipped) / all.length * 100) : 0;
  document.querySelector("#legend .pct").textContent = pct + "%";
  document.querySelector("#topbar .fill-ok").style.width = (all.length ? approved / all.length * 100 : 0) + "%";
  document.querySelector("#topbar .fill-skip").style.width = (all.length ? skipped / all.length * 100 : 0) + "%";
}}
function selectRow(r) {{
  rows().forEach(x => x.classList.toggle("sel", x === r));
  window._sel = r;
}}
function answer(r, which) {{
  const box = r.querySelector("input." + which);
  const other = r.querySelector("input." + (which === "approve" ? "skip" : "approve"));
  box.checked = !box.checked;
  if (box.checked) other.checked = false;
  save(); refresh();
}}
document.addEventListener("change", e => {{
  const r = e.target.closest(".row.cand");
  if (r && e.target.matches("input.approve, input.skip") && e.target.checked) {{
    const other = r.querySelector("input." + (e.target.classList.contains("approve") ? "skip" : "approve"));
    other.checked = false;
    // Answering a row moves on: deselect so it collapses, like the split report.
    if (window._sel === r) selectRow(null);
  }}
  save(); refresh();
}});
document.addEventListener("input", save);
document.addEventListener("click", e => {{
  const r = e.target.closest(".row.cand");
  if (!r) return;
  if (e.target.closest(".head")) selectRow(r);
  else if (!r.classList.contains("sel")) selectRow(r);
}});
// The most recently played player is what the arrow keys scrub.
document.querySelectorAll("audio").forEach(a => {{
  a.addEventListener("play", () => {{
    document.querySelectorAll("audio").forEach(o => {{ if (o !== a && !o.paused) o.pause(); }});
    window._playing = a;
    const r = a.closest(".row.cand");
    if (r) selectRow(r);
  }});
}});
document.addEventListener("keydown", e => {{
  if (e.metaKey || e.ctrlKey || e.altKey) return;
  const t = e.target && e.target.tagName;
  const editing = t === "SELECT" || t === "TEXTAREA"
    || (t === "INPUT" && e.target.type === "text") || (e.target && e.target.isContentEditable);
  if (editing) return;
  const all = rows();
  const cur = window._sel;
  if (e.key === "ArrowDown" || e.key === "ArrowUp") {{
    e.preventDefault();
    const i = cur ? all.indexOf(cur) : -1;
    const next = all[Math.min(all.length - 1, Math.max(0, i < 0 ? 0 : i + (e.key === "ArrowDown" ? 1 : -1)))];
    if (!next) return;
    selectRow(next);
    next.scrollIntoView({{block: "center"}});
    return;
  }}
  if (e.key === "ArrowLeft" || e.key === "ArrowRight") {{
    const a = window._playing;
    if (!a) return;
    e.preventDefault();
    const step = (e.shiftKey ? 5 : 1) * (e.key === "ArrowRight" ? 1 : -1);
    const d = a.duration || Infinity;
    a.currentTime = Math.min(d, Math.max(0, a.currentTime + step));
    return;
  }}
  if (e.key === " ") {{
    const a = window._playing;
    if (!a) return;
    e.preventDefault();
    if (e.target && e.target.blur) e.target.blur();
    return a.paused ? a.play().catch(() => {{}}) : a.pause();
  }}
  const k = e.key.toLowerCase();
  if ((k === "w" || k === "k" || k === "x") && cur) {{
    e.preventDefault();
    answer(cur, k === "w" ? "approve" : "skip");
    const i = all.indexOf(cur);
    const next = all[Math.min(all.length - 1, i + 1)];
    if (next && next !== cur) {{ selectRow(next); next.scrollIntoView({{block: "center"}}); }}
    else selectRow(null);
  }}
}});
restore();
refresh();
document.getElementById("export").onclick = () => {{
  const approved = rows()
    .filter(r => r.querySelector("input.approve").checked)
    .map(r => Object.assign(payloadOf(r), {{
      title: r.querySelector("input.title").value || "Banter",
      song_id: Number(r.querySelector("select.song").value),
      song_title: r.querySelector("select.song").selectedOptions[0]?.textContent,
    }}));
  const blob = new Blob([JSON.stringify(approved, null, 2)], {{type: "application/json"}});
  const a = Object.assign(document.createElement("a"),
    {{href: URL.createObjectURL(blob), download: "approved.json"}});
  a.click();
}};
</script>
""")


def report_to_dict(r):
    d = asdict(r)
    return d


def report_from_dict(d):
    d = dict(d)
    d["candidates"] = [Candidate(**c) for c in d["candidates"]]
    return ShowReport(**d)


def main():
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--out-dir", type=Path, required=True)
    p.add_argument("--dates", default="", help="comma-separated YYYY-MM-DD list")
    p.add_argument("--from-notes", action="store_true",
                   help="scan every show's taper notes for missing banter entries")
    p.add_argument("--rebuild", action="store_true",
                   help="rewrite review.html from the existing report.json")
    args = p.parse_args()
    out_dir = args.out_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / "report.json"

    existing = {}
    if report_path.exists():
        existing = {r["date"]: r for r in json.loads(report_path.read_text())["shows"]}

    if args.rebuild:
        write_review(out_dir, [report_from_dict(d) for d in existing.values()])
        print(f"Review page written to {out_dir / 'review.html'}", file=sys.stderr)
        return

    dates = [d.strip() for d in args.dates.split(",") if d.strip()]
    if args.from_notes:
        dates += dates_from_notes()
    if not dates:
        p.error("give --dates or --from-notes")

    for date in sorted(set(dates)):
        print(f"{date}", file=sys.stderr)
        try:
            report = analyze_show(date, out_dir)
        except Exception as e:  # keep going; the page shows what failed
            report = ShowReport(date=date, venue="", url=f"{SITE_BASE}/{date}", note_hits=[],
                                source_item=None, matched=0, total=0, tried=[],
                                unmatched_files=[], candidates=[], error=f"{type(e).__name__}: {e}")
        existing[date] = report_to_dict(report)
        n = sum(1 for c in report.candidates if c.status == "candidate")
        print(f"  {report.error or f'{report.source_item}: {n} candidate(s)'}", file=sys.stderr)
        report_path.write_text(json.dumps({"shows": list(existing.values())}, indent=2))

    write_review(out_dir, [report_from_dict(d) for d in existing.values()])
    print(f"Review page written to {out_dir / 'review.html'}", file=sys.stderr)


if __name__ == "__main__":
    main()
